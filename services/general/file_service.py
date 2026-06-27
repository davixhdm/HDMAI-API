"""
HDM AI - File Processing Service
Extract text from uploaded files: PDF, images, CSV, Excel, text
"""

import io
import csv
from typing import List, Optional
from loguru import logger


class FileService:
    """Extract text content from uploaded files."""

    SUPPORTED_TYPES = {
        'pdf': 'application/pdf',
        'png': 'image/png',
        'jpg': 'image/jpeg',
        'jpeg': 'image/jpeg',
        'csv': 'text/csv',
        'xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        'xls': 'application/vnd.ms-excel',
        'txt': 'text/plain',
        'md': 'text/markdown',
        'py': 'text/x-python',
        'js': 'text/javascript',
        'json': 'application/json',
    }

    async def extract_text(self, file) -> str:
        """Extract text from a file based on its extension."""
        if not file or not file.filename:
            return ""

        ext = file.filename.split('.')[-1].lower() if '.' in file.filename else ''
        content = await file.read()

        try:
            if ext == 'pdf':
                return await self._extract_pdf(content)
            elif ext in ('png', 'jpg', 'jpeg'):
                return await self._extract_image(content)
            elif ext == 'csv':
                return self._extract_csv(content)
            elif ext in ('xlsx', 'xls'):
                return self._extract_excel(content)
            elif ext in ('txt', 'md', 'py', 'js', 'json', 'yaml', 'yml'):
                return self._extract_text(content)
            else:
                # Try as plain text
                try:
                    return content.decode('utf-8', errors='ignore')[:5000]
                except:
                    return f"[Binary file: {file.filename}]"
        except Exception as e:
            logger.error(f"File extraction failed for {file.filename}: {e}")
            return f"[Could not read file: {file.filename}]"

    async def extract_multiple(self, files: List) -> List[dict]:
        """Extract text from multiple files."""
        results = []
        for file in files:
            if file and file.filename:
                text = await self.extract_text(file)
                results.append({
                    "filename": file.filename,
                    "content": text,
                })
        return results

    async def _extract_pdf(self, content: bytes) -> str:
        """Extract text from PDF."""
        try:
            import pdfplumber
            with pdfplumber.open(io.BytesIO(content)) as pdf:
                text = ""
                for page in pdf.pages[:5]:  # Max 5 pages
                    page_text = page.extract_text()
                    if page_text:
                        text += page_text + "\n"
                return text[:5000] if text else "[PDF - no extractable text]"
        except ImportError:
            return "[PDF extraction unavailable - install pdfplumber]"
        except Exception as e:
            return f"[PDF error: {e}]"

    async def _extract_image(self, content: bytes) -> str:
        """Extract text from image using Gemini Vision."""
        import base64
        try:
            from services.ai_service import ai_service
            b64 = base64.b64encode(content).decode('utf-8')
            result = await ai_service.gemini_vision(
                prompt="Extract all text from this image. Return only the text.",
                image_base64=b64,
            )
            if result.get("success"):
                return result.get("analysis", "")[:3000]
            return "[Image - could not extract text]"
        except Exception as e:
            return f"[Image error: {e}]"

    def _extract_csv(self, content: bytes) -> str:
        """Extract text from CSV."""
        try:
            text = content.decode('utf-8', errors='ignore')
            reader = csv.reader(io.StringIO(text))
            rows = list(reader)
            if not rows:
                return "[Empty CSV]"
            # Format as readable table
            result = f"CSV with {len(rows)} rows, {len(rows[0]) if rows else 0} columns:\n"
            for row in rows[:50]:  # Max 50 rows
                result += " | ".join(row) + "\n"
            return result[:5000]
        except Exception as e:
            return f"[CSV error: {e}]"

    def _extract_excel(self, content: bytes) -> str:
        """Extract text from Excel."""
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
            text = ""
            for sheet_name in wb.sheetnames[:3]:  # Max 3 sheets
                ws = wb[sheet_name]
                text += f"\n--- Sheet: {sheet_name} ---\n"
                for row in ws.iter_rows(values_only=True, max_row=50):
                    text += " | ".join(str(cell) if cell is not None else "" for cell in row) + "\n"
            return text[:5000]
        except ImportError:
            return "[Excel extraction unavailable - install openpyxl]"
        except Exception as e:
            return f"[Excel error: {e}]"

    def _extract_text(self, content: bytes) -> str:
        """Extract plain text."""
        try:
            return content.decode('utf-8', errors='ignore')[:5000]
        except:
            return "[Could not decode file]"


file_service = FileService()